# xlwings

import sys
import time
import random 
import string
from   os              import sep
from   openpyxl        import Workbook , load_workbook # seems like  openpyxl messes up with title's font colors so i won't use ...
from   openpyxl.styles import Font 
from   colour          import Color
from   datetime        import datetime
from   colorhash       import ColorHash
from   untappdscr      import UntappdScraper



filename        = __file__.rsplit(sep, 1)[0] + sep + 'untappd.xlsx' 
untappd         = UntappdScraper((1,2),debug_mode=True)
wb              = load_workbook(filename)
sheet           = wb.active 
color_gradient1 = list(Color("red"   ).range_to(Color("green" ), 50 ))
color_gradient2 = list(Color("yellow").range_to(Color("maroon"), 200))
DATE_FORMAT     = "%d/%m/%Y, %H:%M:%S"


def num2col(num): # https://stackoverflow.com/a/23862195/11465149
    start_index = 1   #  it can start either at 0 or at 1
    letter = ''
    while num > 25 + start_index:   
        letter += chr(65 + int((num-start_index)/26) - 1)
        num = num - (int((num-start_index)/26))*26
    return letter + chr(65 - start_index + (int(num)))


def col2num(col): # https://stackoverflow.com/a/12640614/11465149
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num


def spllit_colrows(s): # https://stackoverflow.com/a/430665/11465149
    head = s.rstrip('0123456789')
    tail = s[len(head):]
    return head, int(tail)


def get_list_table(at, shuffle=False): 
    col, row = spllit_colrows(at)
    coln = col2num(col)
    breweries = []
    while sheet.cell(row,coln).value != None:
        breweries.append(sheet.cell(row,coln).value)
        row += 1
    if shuffle: random.shuffle(breweries)
    return breweries   


def prnt_common(i,j, obj):
    sheet.cell(i, j    ).value = obj.details.stats.total
    sheet.cell(i, j + 1).value = obj.details.stats.unique
    sheet.cell(i, j + 2).value = obj.details.stats.monthly
    sheet.cell(i, j + 3).value = obj.details.stats.you


def print_breweries(_from, at, color='#000000'): # repeating code but nvm for now
    col, row   = spllit_colrows(at)
    coln       = col2num(col)
    cole       = num2col(coln+13)
    breweries  = get_list_table(_from, True) # random choice or something
    date_color = ColorHash(datetime.now()).hex.split('#')[1]
    color      = color.split('#')[1]
    sheet.move_range(f'{col}{row}:{cole}{sheet.max_row}', rows=len(breweries), cols=0)
    for i, brewery in enumerate(breweries,row):
        brewery = untappd.get_brewery(brewery)
        date_time = datetime.now().strftime(DATE_FORMAT)
        sheet.cell(i, coln     ).value = brewery.details.claimed
        sheet.cell(i, coln     ).font  = Font(color='217346' if brewery.details.claimed else '732121')
        sheet.cell(i, coln + 1 ).value = brewery.details.stats.likes
        prnt_common(i, coln + 2, brewery)
        sheet.cell(i, coln + 6 ).value = brewery.details.beer_count
        sheet.cell(i, coln + 7 ).value = brewery.details.rating
        sheet.cell(i, coln + 7 ).font  = Font(color=color_gradient1[int(brewery.details.rating*10)].get_hex_l().split('#')[1])
        sheet.cell(i, coln + 8 ).value = brewery.details.ratings
        sheet.cell(i, coln + 9 ).value = brewery.details.category
        sheet.cell(i, coln + 9 ).font  = Font(color=ColorHash(brewery.details.category).hex.split('#')[1])
        sheet.cell(i, coln + 10).value = brewery.breweryname
        sheet.cell(i, coln + 10).font  = Font(color=ColorHash(brewery.breweryname).hex.split('#')[1])
        sheet.cell(i, coln + 11).value = brewery.name
        sheet.cell(i, coln + 12).value = brewery.details.location
        sheet.cell(i, coln + 12).font  = Font(color=color)
        sheet.cell(i, coln + 13).value = date_time # i know it repeats itself but i don't know how to manage it diferently on excel, so here we are..
        sheet.cell(i, coln + 13).font = Font(color=date_color )


def print_beers(_from, at): # repeating code but nvm for now
    col, row = spllit_colrows(at)
    coln = col2num(col)
    cole = num2col(coln+12)
    beers  = get_list_table(_from, True) # random choice or something
    date_color = ColorHash(datetime.now()).hex.split('#')[1] # unnecessery computation here but nvm, could had been just random or just boolean based, lol
    sheet.move_range(f'{col}{row}:{cole}{sheet.max_row}', rows=len(beers), cols=0)
    for i, beer in enumerate(beers, row):
        beer      = untappd.get_beer(int(beer)) # int lol
        date_time = datetime.now().strftime(DATE_FORMAT)
        sheet.cell(i, coln     ).value = beer.details.discontinued
        sheet.cell(i, coln     ).font  = Font(color='217346' if beer.details.discontinued else '732121')
        prnt_common(i, coln + 1, beer)
        sheet.cell(i, coln + 5 ).value = beer.details.rating
        sheet.cell(i, coln + 5 ).font  = Font(color=color_gradient1[int(beer.details.rating*10)].get_hex_l().split('#')[1]  )
        sheet.cell(i, coln + 6 ).value = beer.details.ratings
        sheet.cell(i, coln + 7 ).value = beer.details.ABV
        sheet.cell(i, coln + 8 ).value = beer.details.IBU
        sheet.cell(i, coln + 8 ).font  = Font(color=color_gradient2[int(beer.details.IBU)].get_hex_l().split('#')[1] if beer.details.IBU else '000000')
        sheet.cell(i, coln + 9 ).value = beer.id
        sheet.cell(i, coln + 9 ).font  = Font(color=ColorHash(beer.id).hex.split('#')[1])
        sheet.cell(i, coln + 10).value = beer.name
        sheet.cell(i, coln + 11).value = beer.brewery.breweryname
        sheet.cell(i, coln + 11).font  = Font(color=ColorHash(beer.brewery.breweryname).hex.split('#')[1])
        sheet.cell(i, coln + 12).value = date_time
        sheet.cell(i, coln + 12).font  = Font(color=date_color)
        

def print_venues(_from, at):
    col, row   = spllit_colrows(at)
    coln       = col2num(col)
    cole       = num2col(coln+10)
    venue      = get_list_table(_from, True) # random choice or something
    date_color = ColorHash(datetime.now()).hex.split('#')[1] # unnecessery computation here but nvm, could had been just random or just boolean based, lol
    sheet.move_range(f'{col}{row}:{cole}{sheet.max_row}', rows=len(venue), cols=0)
    for i, venue in enumerate(venue, row):
        venue = untappd.get_venue(int(venue)) # int lol
        date_time = datetime.now().strftime(DATE_FORMAT)
        sheet.cell(i, coln     ).value = venue.is_verified
        sheet.cell(i, coln     ).font  = Font(color='217346' if venue.is_verified else '732121')
        prnt_common(i, coln + 1, venue)
        sheet.cell(i, coln + 5 ).value = venue.id
        sheet.cell(i, coln + 5 ).font  = Font(color=ColorHash(venue.id).hex.split('#')[1])
        sheet.cell(i, coln + 6 ).value = venue.name
        sheet.cell(i, coln + 7 ).value = venue.details.address_name
        sheet.cell(i, coln + 8 ).value = venue.details.map_url
        sheet.cell(i, coln + 9 ).value = venue.category
        sheet.cell(i, coln + 9 ).font  = Font(color=ColorHash(venue.category).hex.split('#')[1])
        sheet.cell(i, coln + 10).value = date_time
        sheet.cell(i, coln + 10).font  = Font(color=date_color)


def print_top_rated_beers(at, country='', type='', get_picker=False): # col, row , coutry 
    col,row    = spllit_colrows(at)
    coln       = col2num(col)
    cole       = num2col(coln+9)
    beers      = untappd.get_top_rated_beers(country, type, get_picker).items()
    date_time  = datetime.now().strftime(DATE_FORMAT)
    date_color = ColorHash(date_time).hex.split('#')[1]   
    sheet.move_range(f'{col}{row}:{cole}{sheet.max_row}', rows=len(beers), cols=0)
    for i,(_, beer) in enumerate(beers,row):
        sheet.cell(i, coln     ).value = beer.details.discontinued
        sheet.cell(i, coln     ).font  = Font(color='217346' if beer.details.discontinued else '732121')
        sheet.cell(i, coln + 1 ).value = beer.details.rating
        sheet.cell(i, coln + 1 ).font  = Font(color=color_gradient1[int(beer.details.rating*10)].get_hex_l().split('#')[1])
        sheet.cell(i, coln + 2 ).value = beer.details.ratings
        sheet.cell(i, coln + 3 ).value = beer.details.ABV
        sheet.cell(i, coln + 4 ).value = beer.details.IBU
        sheet.cell(i, coln + 4 ).font  = Font(color=color_gradient2[int(beer.details.IBU)].get_hex_l().split('#')[1] if beer.details.IBU else '000000')
        sheet.cell(i, coln + 5 ).value = beer.id
        sheet.cell(i, coln + 5 ).font  = Font(color=ColorHash(beer.id).hex.split('#')[1])
        sheet.cell(i, coln + 6 ).value = beer.name
        sheet.cell(i, coln + 7 ).value = beer.brewery.breweryname
        sheet.cell(i, coln + 7 ).font  = Font(color=ColorHash(beer.brewery.breweryname).hex.split('#')[1])
        sheet.cell(i, coln + 8 ).value = datetime.strptime(beer.details.date_added ,'%m/%d/%y').strftime(' %d/%m/%Y').replace(' 0', '').replace('/0', '/').replace(' ', '') # ignore this shit 
        sheet.cell(i, coln + 9 ).value = date_time
        sheet.cell(i, coln + 9 ).font  = Font(color=date_color)
        

def print_top_rated_breweries(at, country='', type='', color='#217346', get_picker=False): # col, row , coutry 
    col,row    = spllit_colrows(at)
    coln       = col2num(col)
    cole       = num2col(coln+7)
    breweries  = untappd.get_top_rated_breweries(country, type, get_picker).items()
    date_time  = datetime.now().strftime(DATE_FORMAT)
    date_color = ColorHash(date_time).hex.split('#')[1]
    sheet.move_range(f'{col}{row}:{cole}{sheet.max_row}', rows=len(breweries), cols=0)
    for i,(_, brewery) in enumerate(breweries,row):
        sheet.cell(i, coln    ).value = brewery.details.beer_count
        sheet.cell(i, coln + 1).value = brewery.details.rating
        sheet.cell(i, coln + 1).font  = Font(color=color_gradient1[int(brewery.details.rating*10)].get_hex_l().split('#')[1])
        sheet.cell(i, coln + 2).value = brewery.details.ratings
        sheet.cell(i, coln + 3).value = brewery.details.category
        sheet.cell(i, coln + 3).font  = Font(color=ColorHash(brewery.details.category).hex.split('#')[1])
        sheet.cell(i, coln + 4).value = brewery.breweryname
        sheet.cell(i, coln + 4).font  = Font(color=ColorHash(brewery.breweryname).hex.split('#')[1])
        sheet.cell(i, coln + 5).value = brewery.name
        sheet.cell(i, coln + 6).value = brewery.details.location
        sheet.cell(i, coln + 6).font  = Font(color=color.split('#')[1])
        sheet.cell(i, coln + 7).value = date_time # i know it repeats itself but i don't know how to manage it diferently on excel, so here we are.. 
        sheet.cell(i, coln + 7).font  = Font(color=date_color)
    

def main(request):
    if   request == 1: print_breweries          ('A4','AF4') 
    elif request == 2: print_beers              ('B4','AU4')
    elif request == 3: print_venues             ('C4','CE4')
    elif request == 4: print_top_rated_beers    ('BI4')
    elif request == 5: print_top_rated_beers    ('BT4', 'Greece')
    elif request == 6: print_top_rated_breweries('E4' , 'Greece',            color='#2F75B5');  
    elif request == 7: print_top_rated_breweries('N4' , 'Greece', 'micro_brewery', '#2F75B5'); 
    else             : print_top_rated_breweries('W4'                                       );   
    wb.save(filename)

if __name__ == "__main__":
    main(int(sys.argv[1])) # distinctypy has a rng seed don't forget  https://distinctipy.readthedocs.io/en/latest/api.html?highlight=seed#distinctipy.distinctipy.get_colors
